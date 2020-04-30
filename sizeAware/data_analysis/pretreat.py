import xlrd
import sys
import numpy as np
import random
import openpyxl
import os
from xlutils.copy import copy
from sklearn.preprocessing import MinMaxScaler


def getData(filePath):
    file = xlrd.open_workbook(filePath)
    sheet = file.sheet_by_name("data")
    rows = sheet.row_values(1)
    row = sheet.nrows
    for i in range(2, row):
        rows = np.vstack((rows, sheet.row_values(i)))
    a = rows[:, :sheet.ncols-2]
    #a = matrix(np.array(rows).astype(float))
    return a


def saveData(output, data):
    if(os.path.exists(output)):
        os.remove(output)
    wb = openpyxl.Workbook()
    data_sheet = wb.create_sheet(title="data")
    wb.remove(wb["Sheet"])
    rows = data.shape[0]
    cols = data.shape[1]
    for i in range(rows):
        for j in range(cols):
            data_sheet.cell(row=i+1, column=j+1).value = data[i, j]
    wb.save(output)


def normalize(data):
    row = data.shape[0]
    col = data.shape[1]
    # avg = data.sum(axis=0)/row
    # avg_matrix = np.tile(avg, (row, 1))
    # dis = data.max(axis=0)-data.min(axis=0)
    # Data = data-avg_matrix
    # mul = (avg_matrix-data).T.dot(avg_matrix-data)
    # stand = np.sqrt(np.diag(mul)/(row-1))
    # for i in range(row):
    #     for j in range(col-4):
    #         Data[i, j] = Data[i, j]/dis[0,j]
    min_max_scaler = MinMaxScaler()
    x_std = min_max_scaler.fit_transform(data)
    for i in range(row):
        for j in range(col-5, col):
            x_std[i, j] = data[i, j]
    return x_std

def getTrainAndTest(filePath, data):
    file = openpyxl.load_workbook(filePath)
    if len(file.sheetnames)>1:
        test_sheet = file["test"]
        train_sheet = file["train"]
        file.remove(test_sheet)
        file.remove(train_sheet)
        file.save(filePath)
    test_sheet = file.create_sheet(title="test")
    train_sheet = file.create_sheet(title="train")
    test_rows = 0
    train_rows = 0

    rows = data.shape[0]
    cols = data.shape[1]

    for i in range(0, rows):
        if random.randint(0, 4) == 0:
            for j in range(0, cols):
                test_sheet.cell(row=test_rows+1, column=j+1).value=data[i, j]
            test_rows += 1
        else:
            for j in range(0, cols):
                train_sheet.cell(row=train_rows+1, column=j+1).value=data[i, j]
            train_rows += 1
    file.save(filePath)

if __name__ == '__main__':
    data = getData(sys.argv[1])
    data = normalize(data)
    
    saveData(sys.argv[2], data)

    getTrainAndTest(sys.argv[2], data)
