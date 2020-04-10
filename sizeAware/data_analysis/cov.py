import xlrd
import sys
import numpy as np
import random
import openpyxl
from numpy import *
from xlutils.copy import copy


###################################################################################
#####Function:分析各属性与结果的相关性系数                                       #####
#####Input   :命令行输入的参数,加载数据的路径                                    #####

def getData(argv):
    filePath = argv[1]
    file = xlrd.open_workbook(filePath)
    sheet = file.sheet_by_name("data")
    rows = sheet.row_values(0)
    row = sheet.nrows
    for i in range(1, row):
        rows = np.vstack((rows, sheet.row_values(i)))
    a = matrix(np.array(rows).astype(float))
    a = a[:,:sheet.ncols]
    return a


def content(data):
    data = data[:,:data.shape[1]-4]
    row = data.shape[0]
    col = data.shape[1]
    avg = data.sum(axis=0)/row
    avg_matrix = np.tile(avg, (row, 1))

    mul = (avg_matrix-data).T.dot(avg_matrix-data)

    variance = np.diag(mul)/(row-1)

    cov = array(mul[col-1, :]/(row-1))[0]

    related = np.zeros((col))
    for i in range(col):
        related[i] = cov[i]/sqrt(variance[i]*variance[col-1])
    print(related)
    return


def getTrainAndTest(argv, data):
    filePath = argv[1]
    file = openpyxl.load_workbook(filePath)
    if len(file.sheetnames)>1:
        test_sheet = file["test"]
        train_sheet = file["train"]
        file.remove(test_sheet)
        file.remove(train_sheet)
        file.save(filePath)
    test_sheet = file.create_sheet(title="test")
    train_sheet = file.create_sheet(title="train")
    test_rows = 0
    train_rows = 0

    rows = data.shape[0]
    cols = data.shape[1]

    for i in range(0, rows):
        if random.randint(0, 4) == 0:
            for j in range(0, cols):
                test_sheet.cell(row=test_rows+1, column=j+1).value=data[i, j]
            test_rows += 1
        else:
            for j in range(0, cols):
                train_sheet.cell(row=train_rows+1, column=j+1).value=data[i, j]
            train_rows += 1
    file.save(filePath)

def loadData(argv,sheet_name):
    filePath = argv[1]
    file = xlrd.open_workbook(filePath)
    sheet = file.sheet_by_name(sheet_name)
    rows = sheet.row_values(0)
    row = sheet.nrows
    col = sheet.ncols
    for i in range(1, row):
        rows = np.vstack((rows, sheet.row_values(i)))
    a = matrix(np.array(rows).astype(float))
    return a[:,:col-1],a[:,col-1]

if __name__ == '__main__':
    data = getData(sys.argv)
    content(data)
    getTrainAndTest(sys.argv, data)
